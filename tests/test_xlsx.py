from openpyxl import load_workbook
import os.path
# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx():
    xlsx_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../resources/file_example_XLSX_50.xlsx'))
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == "Mara"
    assert workbook.sheetnames == ['Sheet1']

